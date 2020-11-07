import collections

import openpyxl

from . import data

BORDER_OFFSETS = [
    ("top", (0, -1), "north"),
    ("bottom", (0, +1), "south"),
    ("left", (-1, 0), "west"),
    ("right", (+1, 0), "east")
]

def from_spreadsheet(filepath):
    """Load the rooms, layout and items from a spreadsheet
    """
    wb = openpyxl.load_workbook(filepath)
    try:
        rooms = get_rooms(wb)
        layout = get_layout(wb)
        for name, room in rooms.items():
            for direction, neighbour in layout[name].items():
                room.exits[direction] = rooms[neighbour]
        inventory = get_inventory(wb)
    finally:
        wb.close()

    return {
        "rooms" : rooms,
        "inventory" : inventory
    }

def sheet_to_namedtuples(wb, ws_name, cls):
    """Given an openpyxl workbook yield each row as a class
    based on the worksheet name and header row
    """
    ws = wb[ws_name]
    rows = ws.iter_rows(values_only=True)
    headers = [cell.lower() for cell in next(rows)]
    tuplecls = collections.namedtuple("%sRow" % (ws_name), headers)
    for row in rows:
        yield cls.from_namedtuple(tuplecls(*row))

def find_neighbours(cell):
    neighbours = {}
    ws = cell.parent
    x, y = cell.column, cell.row
    for (edge, (dx, dy), direction) in BORDER_OFFSETS:
        border = getattr(cell.border, edge)
        if border.style is None:
            col, row = x + dx, y + dy
            if (row > 0 and col > 0):
                neighbour_name = ws.cell(column=col, row=row).value
                if neighbour_name:
                    neighbours[direction] = neighbour_name

    return neighbours

def get_layout(wb):
    """Return a dictionary mapping each room to its neighbours
    """
    layout = {}
    for row in wb['Layout']:
        for cell in row:
            room_name = cell.value
            if room_name:
                layout[room_name] = find_neighbours(cell)

    return layout

def get_inventory(wb):
    """Return a dictionary mapping inventory names to details"""
    return dict((i.name.lower(), i) for i in sheet_to_namedtuples(wb, "Items", data.Item))

def get_rooms(wb):
    """Return a dictionary mapping room names to details"""
    return dict((r.name, r) for r in sheet_to_namedtuples(wb, "Rooms", data.Room))

